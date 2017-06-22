"""
This module is to compare mwpersistence package with WikiWho.
"""
import csv
import argparse
from os.path import realpath, exists, dirname
from os import mkdir
from time import strftime
from concurrent.futures import ProcessPoolExecutor, as_completed
import deltas
import mwpersistence
from mwreverts.defaults import RADIUS
from deltas.tokenizers.wikitext_split import wikitext_split
from openpyxl import load_workbook
from collections import defaultdict
import json
import traceback
import sys
from pprint import pprint

sys.path.insert(0, dirname(dirname(dirname(realpath(__file__)))))
from wikiwho_api import settings  # dont delete: this overwrites settings in api.utils!
from api.utils import get_latest_revision_data, get_page_data_from_wp_api


def get_token_value(token):
    return token.__repr__().split('Token(')[-1].split(', ')[0][1:-1]


def test_authorship(article_title, token_data, current_tokens_data):
    # prepare token data to test authorship
    correct_rev_id = token_data.get('correct_rev_id')
    token = token_data.get('str')
    context = token_data.get('context')
    sub_token_list = []
    for t in wikitext_split.tokenize(context):
        t_value = get_token_value(t)
        if t_value.replace('\\n', '').replace('\r\n', '\n').replace('\r', '\n').strip():
            # remove white spaces
            sub_token_list.append(t_value)

    n = len(sub_token_list)
    # get list of tokens and revisions of each token of last revision
    token_list = []
    rev_ids = []
    for d in current_tokens_data:
        token_list.append(d['str'])
        rev_ids.append(d['revisions'])
    # check if correct origin rev id and computed origin rev ids are same
    found = 0
    output_dict = {}
    for i in range(len(token_list) - n + 1):
        if sub_token_list == token_list[i:i + n]:
            try:
                token_i = i + sub_token_list.index(token)
                mw_rev_id = rev_ids[token_i][0]  # first one is the first used rev id (origin)
                found = 1
                print(article_title, ' : ', token, found, mw_rev_id == correct_rev_id, mw_rev_id, correct_rev_id)
                if mw_rev_id != correct_rev_id:
                    output_dict[token] = "2) {}: {}: rev id was {}, should be {}".format(article_title,
                                                                                         token,
                                                                                         mw_rev_id,
                                                                                         correct_rev_id)
            except ValueError:
                output_dict[token] = "1) {}: {} -> token does not exist (different tokenization)".\
                    format(article_title, token)
            break
    # print(token, found, len(token_list), len(sub_token_list), sub_token_list)
    if not found:
        output_dict[token] = "3) {}: {} -> token not found".format(article_title, token)
        # print(output_dict)
    return output_dict


def compute_persistence(article_title, article_data, source, output_persistence, output_rev_ids, revert_radius):
    output_dict = {}
    try:
        # wikitext_split is used, default is text_split.
        state = mwpersistence.DiffState(diff_engine=deltas.SegmentMatcher(tokenizer=wikitext_split),
                                        revert_radius=revert_radius)

        d = get_latest_revision_data(article_title=article_title)
        page_id = d['page_id']
        latest_revision_id = d['latest_revision_id']
        if len(article_data) > 1:
            raise NotImplementedError('Processing multi revisions per article is not implemented yet.')

        for rev_id in article_data:
            article_rev_ids = []
            until_revision_id = int(rev_id or latest_revision_id)
            # holds the last revision id which is saved. 0 for new article
            if source == 'api':
                params = {'pageids': page_id, 'action': 'query', 'prop': 'revisions',
                          'rvprop': 'content|ids|timestamp|sha1|comment|flags|user|userid',
                          'rvlimit': 'max', 'format': 'json', 'continue': '', 'rvdir': 'newer',
                          'rvendid': until_revision_id}  # NOTE: rvendid
                for rev_data in get_page_data_from_wp_api(params):
                    current_rev_id = int(rev_data['revid'])
                    text = rev_data.get('*', '').lower()  # NOTE: lower()
                    # process new revision text
                    current_tokens, tokens_added, tokens_removed = state.update(text, revision=current_rev_id)
                    article_rev_ids.append(current_rev_id)
            elif source == 'dumps':
                raise NotImplementedError

            last_rev_tokens_data = []
            for ct in current_tokens:
                ct_value = get_token_value(ct)
                if ct_value.replace('\\n', '').replace('\r\n', '\n').replace('\r', '\n').strip():
                    # remove tokens as white spaces
                    last_rev_tokens_data.append({'str': ct_value,
                                                 'revisions': ct.revisions})

            # output json files for last revision
            json_data = {
                "article_title": article_title,
                "page_id": page_id,
                "revisions": [{current_rev_id: {"tokens": last_rev_tokens_data}}]
            }
            with open(output_persistence, 'w', encoding='utf-8') as f:
                f.write(json.dumps(json_data, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))

            json_data_rev_ids = {
                "article_title": article_title,
                "page_id": page_id,
                "revision_ids": article_rev_ids
            }
            with open(output_rev_ids, 'w', encoding='utf-8') as f:
                f.write(json.dumps(json_data_rev_ids, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))

            # test authorship
            for token_data in article_data[rev_id]:
                if token_data.get('context') and token_data.get('correct_rev_id'):
                    output_dict.update(
                        test_authorship(article_title, token_data, last_rev_tokens_data)
                    )
    except Exception as e:
        print('\n')
        print('=' * 30)
        print(article_title)
        print(traceback.format_exc())
        print('=' * 30)
        print('\n')
    return output_dict


def get_args():
    """
    python mw_persistence.py -m=4 -o='/home/kenan/PycharmProjects/wikiwho_api/tests_ignore/mwpersistence/15'
    python mw_persistence.py -m=6 -o='/home/kenan/PycharmProjects/wikiwho_api/tests_ignore/mwpersistence/9999999999999' -r=9999999999999
python mw_persistence.py -m=4 -p='/home/kenan/PycharmProjects/wikiwho_api/tests_ignore/mwpersistence/random_1000/1000_random_articles.csv' -o='/home/kenan/PycharmProjects/wikiwho_api/tests_ignore/mwpersistence/random_1000/15'
    """
    parser = argparse.ArgumentParser(description='Compute content persistence and token authorship by '
                                                 'using mwpersistence package. This module is created to compare '
                                                 'results of mwpersistence with WikiWho.')

    parser.add_argument('-p', '--articles', help='Path of a csv file that contains list of articles to process. '
                                                 'It should contain page id and rev id (until where persistence is '
                                                 'calculated). Default is gold standard excel file.')
    parser.add_argument('-s', '--revision_source', choices=['api', 'dumps'],
                        help='Source for revision texts. Default is wp api.')
    parser.add_argument('-m', '--max_workers', type=int, help='Default is 16')
    parser.add_argument('-o', '--output_folder', required=True)
    parser.add_argument('-r', '--revert_radius', type=int, help='Default is from mwreverts (15).')
    args = parser.parse_args()
    return args


def main():
    args = get_args()

    output_folder = args.output_folder
    if not exists(output_folder):
        mkdir(output_folder)

    articles_file = args.articles
    articles_dict = defaultdict(dict)
    if articles_file is None:
        # read from gold standard xml file
        articles_file = '{}/{}'.format(dirname(realpath(__file__)), 'test_wikiwho_simple.xlsx')
        wb = load_workbook(filename=articles_file, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue
            if not row[0].value:
                # parse excel file until first empty row
                break
            article_title = row[0].value.replace(' ', '_')
            revision_id = int(row[2].value)
            # this data is used in token authorship test
            if revision_id in articles_dict[article_title]:
                # NOTE: lower text
                articles_dict[article_title][revision_id].append({
                    'str': '{}'.format(row[3].value).lower(),
                    'context': '{}'.format(row[4].value).lower(),
                    'correct_rev_id': int(row[5].value),
                })
            else:
                articles_dict[article_title][revision_id] = [{
                    'str': '{}'.format(row[3].value).lower(),
                    'context': '{}'.format(row[4].value).lower(),
                    'correct_rev_id': int(row[5].value),
                }]
    else:
        # read from input csv file
        with open(articles_file, 'r') as f:
            csv_file = csv.reader(f, delimiter=',')
            next(csv_file, None)  # skip the headers
            for row in csv_file:
                article_title = row[0].replace(' ', '_')
                # no token authorship test
                if len(row) > 1:
                    revision_id = int(row[1])
                    articles_dict[article_title][revision_id] = []
                else:
                    articles_dict[article_title][None] = []

    revision_source = args.revision_source or 'api'
    max_workers = args.max_workers or 16
    revert_radius = args.revert_radius or RADIUS  # default is 15

    print('max_workers:', max_workers, 'len articles:', len(articles_dict), 'revert_radius:', revert_radius)
    print("Start: ", strftime("%Y-%m-%d-%H:%M:%S"))
    output_dict = {}
    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        jobs = {}
        articles_left = len(articles_dict)
        articles_all = len(articles_dict)
        articles_list = list(articles_dict.keys())
        articles_iter = iter(articles_list)
        while articles_left > 0:
            for article_title in articles_iter:
                # print(article_title, articles_dict[article_title], revision_source, revert_radius)
                # articles_left -= 1
                # continue
                output_persistence = '{}/{}.json'.format(output_folder, article_title)
                output_rev_ids = '{}/{}_rev_ids.json'.format(output_folder, article_title)
                if exists(output_persistence) and exists(output_rev_ids):
                    print('skipping {} - already processed, output files exist'.format(article_title))
                    with open(output_persistence) as f:
                        j = json.load(f)
                        for rev_id, tokens in j['revisions'][0].items():
                            rev_id = int(rev_id)
                            last_rev_tokens_data = tokens['tokens']
                            break
                    for token_data in articles_dict[article_title][rev_id]:
                        if token_data.get('context') and token_data.get('correct_rev_id'):
                            output_dict.update(
                                test_authorship(article_title, token_data, last_rev_tokens_data)
                            )
                    articles_left -= 1
                    continue
                job = executor.submit(compute_persistence,
                                      article_title, articles_dict[article_title],
                                      revision_source, output_persistence, output_rev_ids,
                                      revert_radius)
                jobs[job] = article_title
                if len(jobs) == max_workers:  # limit # jobs with max_workers
                    break

            for job in as_completed(jobs):
                articles_left -= 1
                article_title_ = jobs[job]
                data = job.result()
                output_dict.update({article_title_: data})

                del jobs[job]
                sys.stdout.write('\rArticles left: {} - Done: {:.3f}%\n'.
                                 format(articles_left, ((articles_all - articles_left) * 100) / articles_all))
                break  # to add a new job, if there is any
    pprint(output_dict)
    print("\nDone: ", strftime("%Y-%m-%d-%H:%M:%S"))


if __name__ == '__main__':
    main()
