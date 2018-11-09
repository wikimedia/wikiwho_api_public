# -*- coding: utf-8 -*-
"""
These are integration tests to test behaviour of wikiwho algorithm.

To run only test_authorship test calls:
py.test test_wikiwho_simple.py::TestWikiwho::test_authorship
To run only test_json_output test calls:
py.test test_wikiwho_simple.py::TestWikiwho::test_json_output  --lines=2,32 -n=3
To run tests on specific lines of input excel:
py.test test_wikiwho_simple.py --lines=3,9
py.test test_wikiwho_simple.py::TestWikiwho::test_authorship --lines=3,9
"""
from __future__ import absolute_import
from __future__ import unicode_literals

import sys
import pytest
import json
import filecmp
import io
from builtins import range
from os import mkdir
from os.path import realpath, exists, dirname, join

from openpyxl import load_workbook
from mwxml import Dump
from mwtypes.files import reader

from django.conf import settings

from api.handler import WPHandler
from wikiwho.tests.utils import article_zips
from WikiWho.utils import split_into_tokens, iter_rev_tokens
from api.views import WikiwhoView
from django.utils.dateparse import parse_datetime

# TODO use django's default test tool: https://docs.djangoproject.com/en/1.10/topics/testing/


def pytest_generate_tests(metafunc):
    """
    Generates tests dynamically according to lines command option.
    :param metafunc:
    :return:
    """
    # print(metafunc)  # http://docs.pytest.org/en/latest/parametrize.html#the-metafunc-object
    # print(metafunc.config.option.lines)
    articles = set()
    lines = metafunc.config.option.lines
    lines = [] if lines == 'all' else [int(l) for l in lines.split(',')]
    if 'article_title' in metafunc.fixturenames:
        input_file = '{}/{}'.format(dirname(realpath(__file__)), 'test_wikiwho_simple.xlsx')
        wb = load_workbook(filename=input_file, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows()):
            if i == 0 or lines and i+1 not in lines:
                continue
            if not row[0].value:
                # parse excel file until first empty row
                break
            article_title = row[0].value.replace(' ', '_')
            if 'token' in metafunc.fixturenames:
                # test_authorship
                funcargs = {
                    'article_title': article_title,
                    'revision_id': int(row[2].value),
                    'token': '{}'.format(row[3].value).lower(),
                    'context': '{}'.format(row[4].value).lower(),
                    'correct_rev_id': int(row[5].value),
                }
                metafunc.addcall(funcargs=funcargs)
            elif article_title not in articles and 'revision_id_start' in metafunc.fixturenames:
                # test_continue_logic
                articles.add(article_title)
                funcargs = {
                    'article_title': article_title,
                    'revision_id_start': int(row[5].value),
                    'revision_id_end': int(row[2].value),
                }
                metafunc.addcall(funcargs=funcargs)
            elif article_title not in articles:
                # test_json_output(_xml)
                articles.add(article_title)
                funcargs = {
                    'article_title': article_title,
                    'revision_id': int(row[2].value),
                }
                metafunc.addcall(funcargs=funcargs)


def _test_json(wp, temp_folder, article_title, extended_test=True, from_db=False):
    test_json_folder = '{}/tests_ignore/jsons/after_wikipedia_editor_problem'.\
        format(dirname(dirname(dirname(realpath(__file__)))))

    v = WikiwhoView()
    # create json with rev and editor ids
    revision_json = v.get_revision_content(wp, {'str', 'o_rev_id', 'editor'}, from_db=from_db, with_token_ids=False)
    # compare jsons with rev and editor ids
    json_file_path = '{}/{}_ri_ai.json'.format(temp_folder, article_title)
    with io.open(json_file_path, 'w', encoding='utf-8') as f:
        f.write(json.dumps(revision_json, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))
    is_content_same_1 = filecmp.cmp(json_file_path, '{}/{}_ri_ai.json'.format(test_json_folder, article_title))

    # create json with token ids
    revision_json = v.get_revision_content(wp, {'str', 'token_id'}, from_db=from_db, with_token_ids=True)
    # check if all token ids are unique
    for _, rev in revision_json['revisions'][0].items():
        token_ids = [t['token_id'] for t in rev['tokens']]
        assert len(token_ids) == len(set(token_ids)), "{}: there are duplicated token ids".format(article_title)
        break
    # compare jsons with token ids
    json_file_path = '{}/{}_ti.json'.format(temp_folder, article_title)
    with io.open(json_file_path, 'w', encoding='utf-8') as f:
        f.write(json.dumps(revision_json, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))
    is_content_same_2 = filecmp.cmp(json_file_path, '{}/{}_ti.json'.format(test_json_folder, article_title))

    # create json with in/outbounds
    in_out_test_len = True
    in_out_test_spam = True
    in_out_test_ts = True
    # last_used_test_spam = True
    revision_json = v.get_revision_content(wp, {'str', 'in', 'out'}, from_db=from_db, with_token_ids=False)
    for i, ri in enumerate(wp.revision_ids):
        for t in revision_json['revisions'][i][ri]['tokens']:
            in_out_test_len = 0 <= len(t['out']) - len(t['in']) <= 1
            for r in t['out']:
                if r in wp.wikiwho.spam_ids:
                    in_out_test_spam = False
            for r in t['in']:
                if r in wp.wikiwho.spam_ids:
                    in_out_test_spam = False
            for o, i_ in zip(t['out'], t['in']):
                ts_diff = (parse_datetime(wp.wikiwho.revisions[i_].timestamp) -
                           parse_datetime(wp.wikiwho.revisions[o].timestamp)).total_seconds()
                if ts_diff < 0:
                    in_out_test_ts = False
            if len(t['out']) > len(t['in']) and t['in']:
                ts_diff = (parse_datetime(wp.wikiwho.revisions[t['out'][-1]].timestamp) -
                           parse_datetime(wp.wikiwho.revisions[t['in'][-1]].timestamp)).total_seconds()
                if ts_diff < 0:
                    in_out_test_ts = False
            # last_used_test_spam = not (t['last_used'] in wp.wikiwho.spam_ids)
            if not in_out_test_len or not in_out_test_spam or not in_out_test_ts:
                # print(t['str'], len(t['out']), len(t['in']), t['out'], t['in'])
                break

    # test spams
    file_path_spams = '{}/{}_spams.txt'.format(temp_folder, article_title)
    with io.open(file_path_spams, 'w', encoding='utf-8') as f:
        for spam_id in wp.wikiwho.spam_ids:
            f.write('{}\n'.format(spam_id))
    with open('{}/{}_spams.txt'.format(test_json_folder, article_title), 'r') as f:
        correct_spam_ids = f.read().splitlines()
    # pickles from xml dumps has longer rev ids and spams. so we have to compare them until what is last from api.
    spam_ids_same = True
    for i, j in zip(correct_spam_ids, wp.wikiwho.spam_ids):
        if str(i) != str(j):
            spam_ids_same = False
            break

    # create article rev ids json
    rev_ids_json = v.get_revision_ids(wp, {'rev_id', 'editor', 'timestamp'}, from_db=from_db)
    json_file_path = '{}/{}_rev_ids.json'.format(temp_folder, article_title)
    with io.open(json_file_path, 'w', encoding='utf-8') as f:
        f.write(json.dumps(rev_ids_json, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))
    # compare article rev ids jsons
    with open('{}/{}_rev_ids.json'.format(test_json_folder, article_title), 'r') as f:
        correct_rev_ids = json.load(f)
    # pickles from xml dumps has longer rev ids and spams. so we have to compare them until what is last from api.
    rev_ids_same = True
    for i, j in zip(correct_rev_ids['revisions'], rev_ids_json['revisions']):
        if i != j:
            rev_ids_same = False
            break

    if extended_test:
        # compare jsons with in/outbounds
        json_file_path = '{}/{}_io.json'.format(temp_folder, article_title)
        with io.open(json_file_path, 'w', encoding='utf-8') as f:
            f.write(json.dumps(revision_json, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))
        is_content_same_3 = filecmp.cmp(json_file_path, '{}/{}_io.json'.format(test_json_folder, article_title))

        # TODO get deleted and all content of a specific revision, so that we can test it for xml dumps
        # create deleted content json with threshold 0
        deleted_tokens_json = v.get_deleted_content(wp, ['str', 'o_rev_id', 'editor', 'token_id', 'in', 'out', 0], from_db=from_db)
        json_file_path = '{}/{}_deleted_content.json'.format(temp_folder, article_title)
        with io.open(json_file_path, 'w', encoding='utf-8') as f:
            f.write(json.dumps(deleted_tokens_json, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))
        # compare deleted content jsons
        is_content_same_4 = filecmp.cmp(json_file_path, '{}/{}_deleted_content.json'.format(test_json_folder, article_title))
        # create all content json with threshold 0
        all_tokens_json = v.get_all_content(wp, ['str', 'o_rev_id', 'editor', 'token_id', 'in', 'out', 0], from_db=from_db)
        json_file_path = '{}/{}_all_content.json'.format(temp_folder, article_title)
        with io.open(json_file_path, 'w', encoding='utf-8') as f:
            f.write(json.dumps(all_tokens_json, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))
        # compare all content jsons
        is_content_same_5 = filecmp.cmp(json_file_path, '{}/{}_all_content.json'.format(test_json_folder, article_title))

    # print(wp.wikiwho.spam_ids)
    assert is_content_same_1, "{}: 'json with ri and ai doesn't match".format(article_title)
    assert is_content_same_2, "{}: json with ti doesn't match".format(article_title)
    assert in_out_test_len and in_out_test_spam and in_out_test_ts, "len: {}, spam: {}, ts: {}".format(
        in_out_test_len,  in_out_test_spam, in_out_test_ts)
    if extended_test:
        assert is_content_same_3, "{}: 'json with in/outbounds doesn't match".format(article_title)
        assert is_content_same_4, "{}: 'deleted content json doesn't match".format(article_title)
        assert is_content_same_5, "{}: 'all content json doesn't match".format(article_title)
    # assert last_used_test_spam, 'last used in spam'
    assert rev_ids_same, "{}: 'article rev ids json doesn't match".format(article_title)
    assert spam_ids_same, "{}: spam ids don't match".format(article_title)


def _test_json_from_xml(temp_folder, article_title, revision_id, save_tables=()):
    # read from xml dumps (7z)
    xml_file_path = '{}/tests_ignore/xml_dumps/{}'.format(dirname(dirname(dirname(realpath(__file__)))),
                                                          article_zips[article_title])
    if not exists(xml_file_path):
        # server
        xml_file_path = '/home/nuser/pdisk/xmls_7z/batch_all/{}'.format(article_zips[article_title])
    dump = Dump.from_file(reader(xml_file_path))
    title_matched = False
    for page in dump:
        if page.title.replace(' ', '_') == article_title:
            title_matched = True
            with WPHandler(article_title, page.id, pickle_folder=temp_folder['pickle_folder'], is_xml=True,
                           save_tables=save_tables) as wp:
                wp.handle_from_xml_dump(page)
                wp.revision_ids = [revision_id]
            break

    assert title_matched, '{}--{}'.format(article_title, page.title)
    _test_json(wp, temp_folder['test_json_output_xml'], article_title, extended_test=False)


class TestWikiwho:

    @classmethod
    def setup_class(cls):
        """ setup any state specific to the execution of the given class (which
        usually contains tests).
        """
        settings.TESTING = True
        # sys.path.append(dirname(realpath(__file__)))

    @pytest.fixture(scope='session')
    def temp_folder(self, tmpdir_factory):
        """
        Creates temporary folder for whole test session.
        :param tmpdir_factory:
        :return:
        """
        temp_folder_dict = {}
        if int(sys.version[0]) > 2:
            tmp = 'tmp_test_3'
        else:
            tmp = 'tmp_test'
        tmp = join(dirname(realpath(__file__)), tmp)
        if not exists(tmp):
            mkdir(tmp)
        temp_folder_dict['pickle_folder'] = tmp
        tmp_test_authorship = '{}/test_authorship'.format(tmp)
        if not exists(tmp_test_authorship):
            mkdir(tmp_test_authorship)
        tmp_test_json_output = '{}/test_json_output'.format(tmp)
        temp_folder_dict['tmp_test_json_output'] = tmp_test_json_output
        if not exists(tmp_test_json_output):
            mkdir(tmp_test_json_output)
        test_json_output_xml = '{}/test_json_output_xml'.format(tmp)
        if not exists(test_json_output_xml):
            mkdir(test_json_output_xml)
        temp_folder_dict['test_json_output_xml'] = test_json_output_xml
        tmp_test_continue_logic = '{}/test_continue_logic'.format(tmp)
        if not exists(tmp_test_continue_logic):
            mkdir(tmp_test_continue_logic)
        temp_folder_dict['tmp_test_continue_logic'] = tmp_test_continue_logic
        return temp_folder_dict

    def test_json_output(self, temp_folder, article_title, revision_id):
        """
        Tests json outputs of given revisions of articles in gold standard. Revision data is taken from wp api and
        processed data is stored in pickles.
        If there are unexpected differences in jsons, authorship of each token in gold standard should be
        checked by running test_authorship or manually.
        :param revision_id: Revision id where authorship of token in gold standard is tested.
        """
        with WPHandler(article_title, pickle_folder=temp_folder['pickle_folder']) as wp:
            wp.handle([revision_id], is_api_call=False)
        _test_json(wp, temp_folder['tmp_test_json_output'], article_title)

    def test_json_output_xml(self, temp_folder, article_title, revision_id):
        """
        Tests json outputs of given revisions of articles in gold standard. Revision data is taken from wp xml dumps and
        processed data is stored in pickles.
        If there are unexpected differences in jsons, authorship of each token in gold standard should be
        checked by running test_authorship or manually.
        The xml file dumps taken from here: https://dumps.wikimedia.org/enwiki/20161101/
        :param revision_id: Revision id where authorship of token in gold standard is tested.
        """
        _test_json_from_xml(temp_folder, article_title, revision_id, save_tables=())

    def test_continue_logic(self, temp_folder, article_title, revision_id_start, revision_id_end):
        """
        :param revision_id_start: Correct revision id of first token of article in gold standard.
        :param revision_id_end: Revision id where authorship of token in gold standard is tested.
        """
        # first create article and revisions until revision_id_start
        with WPHandler(article_title, pickle_folder=temp_folder['pickle_folder']) as wp:
            wp.handle([revision_id_start], is_api_call=False)
        assert wp.already_exists is False

        # continue creating revisions of this article until revision_id_end
        with WPHandler(article_title, pickle_folder=temp_folder['pickle_folder']) as wp:
            wp.handle([revision_id_end], is_api_call=False)
        _test_json(wp, temp_folder['tmp_test_continue_logic'], article_title)
        assert wp.already_exists is True

    def test_finger_lakes(self, temp_folder):
        """
        Tests in/outbounds through finger lake dummy article.
        """
        data = {
            0: {
                'used': [[0], [0], [0], [0], [0], [0], [0], [0]],
                'in': [[], [], [], [], [], [], [], []],
                'out': [[], [], [], [], [], [], [], []],
            },
            1: {
                'used': [[0, 1], [1], [0, 1], [0, 1], [0, 1], [1], [0, 1], [0, 1], [1], [1], [1], [1], [1], [1]],
                'in': [[], [], [], [], [], [], [], [], [], [], [], [], [], []],
                'out': [[], [], [], [], [], [], [], [], [], [], [], [], [], []],
            },
            2: {
                'used': [[0, 1, 2], [1, 2], [0, 1, 2], [0, 1, 2], [0, 1, 2], [1, 2], [0, 1, 2], [0, 1, 2]],
                'in': [[], [], [], [], [], [], [], []],
                'out': [[], [], [], [], [], [], [], []],
            },
            3: {
                'used': [[0, 1, 2, 3], [1, 2, 3], [0, 1, 2, 3], [0, 1, 2, 3], [0, 1, 2, 3], [1, 2, 3], [0, 1, 2, 3], [0, 1, 2, 3], [1, 3], [1, 3], [1, 3], [1, 3], [1, 3], [1, 3]],
                'in': [[], [], [], [], [], [], [], [], [3], [3], [3], [3], [3], [3]],
                'out': [[], [], [], [], [], [], [], [], [2], [2], [2], [2], [2], [2]],
            },
            4: {
                'used': [[0, 1, 2, 3, 4], [1, 2, 3, 4], [0, 1, 2, 3, 4], [0, 1, 2, 3, 4], [0, 1, 2, 3, 4], [1, 2, 3, 4], [0, 1, 2, 3, 4], [0, 1, 2, 3, 4], [1, 3, 4], [1, 3, 4], [1, 3, 4], [1, 3, 4], [4], [4]],
                'in': [[], [], [], [], [], [], [], [], [3], [3], [3], [3], [], []],
                'out': [[], [], [], [], [], [], [], [], [2], [2], [2], [2], [], []],
            },
            5: {
                'used': [[0, 1, 2, 3, 4, 5], [1, 2, 3, 4, 5], [5], [5], [0, 1, 2, 3, 4, 5], [1, 2, 3, 4, 5], [0, 1, 2, 3, 4, 5], [0, 1, 2, 3, 4, 5], [1, 3, 4, 5], [1, 3, 4, 5], [1, 3, 4, 5], [1, 3, 4, 5], [4, 5], [4, 5]],
                'in': [[], [], [], [], [], [], [], [], [3], [3], [3], [3], [], []],
                'out': [[], [], [], [], [], [], [], [], [2], [2], [2], [2], [], []],
            },
            6: {
                'used': [[0, 1, 2, 3, 4, 5, 6], [1, 2, 3, 4, 5, 6], [0, 1, 2, 3, 4, 6], [0, 1, 2, 3, 4, 6], [0, 1, 2, 3, 4, 5, 6], [1, 2, 3, 4, 5, 6], [0, 1, 2, 3, 4, 5, 6], [0, 1, 2, 3, 4, 5, 6], [1, 3, 4, 5, 6], [1, 3, 4, 5, 6], [1, 3, 4, 5, 6], [1, 3, 4, 5, 6], [4, 5, 6], [4, 5, 6]],
                'in': [[], [], [6], [6], [], [], [], [], [3], [3], [3], [3], [], []],
                'out': [[], [], [5], [5], [], [], [], [], [2], [2], [2], [2], [], []],
            },
            27: {
                'used': [[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [0, 1, 2, 3, 4, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [0, 1, 2, 3, 4, 6, 7, 8, 9, 10, 12, 15, 17, 19, 20, 21, 22, 23, 24, 25, 26, 27], [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27], [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [26, 27], [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [1, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [1, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 19, 20, 21, 22, 23, 24, 25, 26, 27], [1, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [1, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [25, 26, 27], [25, 26, 27], [25, 26, 27], [4, 5, 6, 7, 8, 9, 10, 12, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [7, 8, 9, 10, 12, 13, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [7, 8, 9, 10, 12, 13, 15, 17, 19, 20, 21, 22, 23, 24, 25, 26, 27], [8, 9, 10, 12, 13, 15, 17, 18, 19, 22, 23, 24, 25, 26, 27], [9, 10, 12, 13, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [8, 9, 10, 12, 13, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27], [27], [10, 12, 13, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27]],
                'in': [[12, 15, 17], [12, 15, 17], [6, 12, 15, 17], [6, 12, 15, 17, 19], [12, 15, 17, 21], [12, 15, 17], [ ], [12, 15, 17], [12, 15, 17], [3, 12, 15, 17], [3, 12, 15, 17, 19], [3, 12, 15, 17], [3, 12, 15, 17], [], [], [], [12, 15, 17], [12, 15, 17], [12, 15, 17, 19], [12, 15, 17, 22], [12, 15, 17], [12, 15, 17], [], [12, 15, 17]],
                'out': [[11, 13, 16], [11, 13, 16], [5, 11, 13, 16], [5, 11, 13, 16, 18], [11, 13, 16, 20], [11, 13, 16], [ ], [11, 13, 16], [11, 13, 16], [2, 11, 13, 16], [2, 11, 13, 16, 18], [2, 11, 13, 16], [2, 11, 13, 16], [], [], [], [11, 13, 16], [11, 14, 16], [11, 14, 16, 18], [11, 14, 16, 20], [11, 14, 16], [11, 14, 16], [], [11, 14, 16]],
            }
        }
        article_title = 'Finger_Lakes'
        tests = dirname(realpath(__file__))
        # create the pickle file
        from django.core import management
        management.call_command('xml_to_pickle', *['--output', tests])
        # load the pickle and compare last rev id, in and outs with data dict
        with WPHandler(article_title, pickle_folder=tests, save_tables=('article', 'revision', 'token', )) as wp:
            for rev_id in wp.wikiwho.ordered_revisions:
                if rev_id not in data.keys():
                    continue
                i = 0
                rev = wp.wikiwho.revisions[rev_id]

                for word in iter_rev_tokens(rev):
                    last_rev_id = [x for x in word.last_rev_id if x <= rev_id]
                    inbound = [x for x in word.inbound if x <= rev_id]
                    outbound = [x for x in word.outbound if x <= rev_id]
                    assert last_rev_id == data[rev_id]['used'][i], 'last_rev_id does not match, rev id: {} - {}'.format(rev_id, i)
                    assert inbound == data[rev_id]['in'][i], 'inbound does not match, rev id: {} - {}'.format(rev_id, i)
                    assert outbound == data[rev_id]['out'][i], 'outbound does not match, rev id: {} - {}'.format(rev_id, i)
                    i += 1

    def test_authorship(self, temp_folder, article_title, revision_id, token, context, correct_rev_id):
        sub_token_list = split_into_tokens(context)
        n = len(sub_token_list)

        pickle_folder = temp_folder['pickle_folder']
        with WPHandler(article_title, pickle_folder=pickle_folder) as wp:
            wp.handle([revision_id], is_api_call=False)

        token_list, origin_rev_ids = wp.wikiwho.get_revision_text(revision_id)
        found = 0
        # print(wp.wikiwho.spam)
        print(token, context)
        print(sub_token_list)
        for i in range(len(token_list) - n + 1):
            if sub_token_list == token_list[i:i + n]:
                token_i = i + sub_token_list.index(token)
                ww_rev_id = origin_rev_ids[token_i]
                found = 1
                assert ww_rev_id == correct_rev_id, "1){}: {}: rev id was {}, should be {}".format(article_title,
                                                                                                   token,
                                                                                                   ww_rev_id,
                                                                                                   correct_rev_id)
                break
        assert found, "{}: {} -> token not found".format(article_title, token)

    @classmethod
    def teardown_class(cls):
        """ teardown any state that was previously setup with a call to
        setup_class.
        """
        # if sys.version[0] > 2:
        #     tmp = 'tmp_test_3'
        # else:
        #     tmp = 'tmp_test'
        # if exists(tmp):
        #     os.removedirs(tmp)
