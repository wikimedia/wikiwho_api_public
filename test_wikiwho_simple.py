# -*- coding: utf-8 -*-
"""
To run only test_authorship test calls:
py.test test_wikiwho_simple.py::TestWikiwho::test_authorship
To run only test_json_output test calls:
py.test test_wikiwho_simple.py::TestWikiwho::test_json_output
To run tests on specific lines of input excel:
py.test test_wikiwho_simple.py --lines=3,9
py.test test_wikiwho_simple.py::TestWikiwho::test_authorship --lines=3,9
"""
from __future__ import absolute_import
from __future__ import unicode_literals
from builtins import range
from openpyxl import load_workbook
import os
import sys
import pytest
import json
import filecmp
import io

from handler import WPHandler
from structures.Text import splitIntoWords
# TODO tests for structures: splitIntoWords ...


def pytest_generate_tests(metafunc):
    """
    Generates tests dynamically according to lines command option.
    :param metafunc:
    :return:
    """
    # print(metafunc)  # http://docs.pytest.org/en/latest/parametrize.html#the-metafunc-object
    # print(metafunc.config.option.lines)
    articles = set()
    lines = metafunc.config.option.lines
    lines = [] if lines == 'all' else [int(l) for l in lines.split(',')]
    # file_or_dir = metafunc.config.option.file_or_dir  # TODO
    if "article_name" in metafunc.fixturenames:
        wb = load_workbook(filename='test_wikiwho_simple.xlsx', data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows()):
            if i == 0 or lines and i+1 not in lines:
                continue
            if not row[0].value:
                break
            article_name = row[0].value.replace(' ', '_')
            if 'token' in metafunc.fixturenames:
                funcargs = {
                    'article_name': article_name,
                    'revision_ids': [int(row[2].value)],
                    'token': '{}'.format(row[3].value).lower(),
                    'context': '{}'.format(row[4].value).lower(),
                    'correct_rev_id': int(row[5].value),
                }
                metafunc.addcall(funcargs=funcargs)
            elif article_name not in articles:
                articles.add(article_name)
                funcargs = {
                    'article_name': article_name,
                    'revision_ids': [int(row[2].value)],
                }
                metafunc.addcall(funcargs=funcargs)


class TestWikiwho:
    @classmethod
    def setup_class(cls):
        """ setup any state specific to the execution of the given class (which
        usually contains tests).
        """

    @pytest.fixture(scope='session')
    def temp_folder(self, tmpdir_factory):
        """
        Creates temporary folder for whole test session.
        :param tmpdir_factory:
        :return:
        """
        if int(sys.version[0]) > 2:
            tmp = 'tmp_test_3'
        else:
            tmp = 'tmp_test'
        if not os.path.exists(tmp):
            os.mkdir(tmp)
        return tmp
        # in os' tmp dir
        # return tmpdir_factory.getbasetemp()

    def test_authorship(self, temp_folder, article_name, revision_ids, token, context, correct_rev_id):
        sub_text = splitIntoWords(context)
        with WPHandler(article_name, temp_folder) as wp:
            wp.handle(revision_ids, 'json')
        text, authors = wp.wikiwho.get_revision_text(revision_ids[0])
        n = len(sub_text)
        found = 0
        # print(wp.wikiwho.spam)
        print(token, context)
        print(sub_text)
        for i in range(len(text) - n + 1):
            if sub_text == text[i:i + n]:
                token_i = i + sub_text.index(token)
                ww_rev_id = authors[token_i]
                found = 1
                assert ww_rev_id == correct_rev_id, "{}: {}: rev id was {}, should be {}".format(article_name,
                                                                                                 token,
                                                                                                 ww_rev_id,
                                                                                                 correct_rev_id)
                break
        assert found, "{}: {} -> token not found".format(article_name, token)

    def test_json_output(self, temp_folder, article_name, revision_ids):
        with WPHandler(article_name, temp_folder) as wp:
            wp.handle(revision_ids, 'json')
        response = wp.wikiwho.print_revision(wp.revision_ids, {}, return_response=True)
        json_file_path = '{}/{}.json'.format(temp_folder, article_name)
        test_json_file_path = 'test_jsons/{}.json'.format(article_name)
        with io.open(json_file_path, 'w', encoding='utf-8') as f:
            f.write(json.dumps(response, indent=4, separators=(',', ': '), sort_keys=True, ensure_ascii=False))
        is_content_same = filecmp.cmp(json_file_path, test_json_file_path)
        assert is_content_same

    @classmethod
    def teardown_class(cls):
        """ teardown any state that was previously setup with a call to
        setup_class.
        """
        # if sys.version[0] > 2:
        #     tmp = 'tmp_test_3'
        # else:
        #     tmp = 'tmp_test'
        # if os.path.exists(tmp):
        #     os.removedirs(tmp)
